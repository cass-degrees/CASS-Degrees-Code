import csv
from io import TextIOWrapper
import openpyxl

from api.models import CourseModel, SubplanModel
from django.contrib.auth.decorators import login_required
from django.shortcuts import render


# Submit file with courses or subplans with the following formats:
# Note that column order does not matter, as long as data corresponds to the order of the first row.

# Courses:
# code%year%name%units%offeredSem1%offeredSem2
# ARTS1001%2019%Introduction to Arts%6%True%False
# ...

# Subplans:
# code%year%name%units%planType
# ARTI-SPEC%2016%Artificial Intelligence%24%SPEC
# ...

# Support is available for excel sheets in the same rules as above (no % needed, but put each word between
# the % sign in a new cell in that row).

# Support is available for CASS' custom teaching plan excel file.


@login_required
def bulk_data_upload(request):
    context = dict()
    context['upload_type'] = ['Courses', 'Subplans']
    content_type = request.GET.get('type')

    if content_type in context['upload_type']:
        context['current_tab'] = content_type

    if request.method == 'POST':
        base_model_url = request.build_absolute_uri('/api/model/')

        # Open file in text mode:
        # https://stackoverflow.com/questions/16243023/how-to-resolve-iterator-should-return-strings-not-bytes
        raw_uploaded_file = request.FILES['uploaded_file']
        uploaded_file = TextIOWrapper(raw_uploaded_file, encoding=request.encoding)

        # First row contains the column type headings (code, name etc). We can't add them to the db.
        first_row_checked = False

        # Check if any errors or successes appear when uploading the files.
        # Used for determining type of message to show to the user on the progress of their file upload.
        any_error = False
        any_success = False
        failed_to_upload = []
        correctly_uploaded = []

        # If the uploaded file was an excel sheet, convert it to a format that is the same as the output of when
        # the percent separated value files are processed by the else statement below.
        # This is done so that the code below the if else statement here works for both types of files.
        if uploaded_file.name[-4:] == "xlsx" or uploaded_file.name[-3:] == "xls":
            # Used https://www.pythoncircle.com/post/591/how-to-upload-and-process-the-excel-file-in-django/
            # to help me read the excel files.
            excel_file = openpyxl.load_workbook(raw_uploaded_file)
            sheet = excel_file["Sheet1"]

            # This is not the best way to make sure
            cass_course_custom_format = (list(sheet.iter_rows())[0][0].value is None)

            uploaded_file = list()

            # If the uploaded excel sheet is in format of what CASS already has, do special processing.
            # Change this if statement when the checkbox is implemented.
            if cass_course_custom_format:
                year_range_found = False
                columns_set = False
                col_index = dict()
                col_counter = 0

                start_year = int()
                end_year = int()

                # Add columns that are matching the custom format designed in this script, detailed in line 10 - 26.
                uploaded_file.append(["code", "year", "name", "units", "offeredSem1", "offeredSem2"])
                for row in sheet.iter_rows():

                    sem_offer_value = ""
                    if columns_set:
                        sem_offer_value = row[col_index["Semesters"]].value

                    # This is the initialisation phase,
                    # where the file reader determines the year ranges and column positions.
                    if not year_range_found or not columns_set:
                        for cell in row:
                            # Split the string in the first line of the excel sheet to get start and end years.
                            if not year_range_found:
                                if cell.value is not None:
                                    stripped_title = cell.value.split()

                                    # Only extract years if we found the right cell
                                    if " ".join(stripped_title[0:5]) == "CASS 3 Year Teaching Plan:":
                                        start_year = int(stripped_title[-3])
                                        end_year = int(stripped_title[-1])

                                        year_range_found = True

                            # Once the year range is set, then determine the column positions of the excel sheet.
                            elif not columns_set:
                                if cell.value is not None:
                                    # Find and store the index of all the column positions
                                    col_index[cell.value] = col_counter
                                    col_counter += 1

                                # If every column has been indexed, then mark columns_set as true.
                                if col_counter + 1 == len(row):
                                    columns_set = True

                        # If years are not determined in the first row, the excel file is not in the desirable format.
                        if not year_range_found:
                            any_error = True
                            failed_to_upload.append("Unknown File Format")
                            break

                    # Once the year range and the column positions are set,
                    # check the offerings specified in the 'Semester' column and process accordingly.
                    else:
                        # Right now, courses with offering type 'other' and 'sessions' are not supported,
                        # so add them to the list of 'failed' courses and move on to the next course.
                        if sem_offer_value == "Other" or sem_offer_value == "Offered only in sessions":
                            skipped_course = "{} - {} - Unknown/Unsupported course offering".format(
                                row[col_index["Course Code"]].value,
                                row[col_index["Course Title"]].value)

                            failed_to_upload.append(skipped_course)
                            any_error = True
                            continue

                        sem_offer_args = sem_offer_value.split()

                        for year in range(start_year, end_year + 1):
                            s1_offer = False
                            s2_offer = False

                            # Skip for this year if the semester offer states it is not offered in odd/even years.
                            if sem_offer_args[0] == "Even":
                                if year % 2 == 1:
                                    continue
                            elif sem_offer_args[0] == "Odd":
                                if year % 2 == 0:
                                    continue

                            # Set the boolean values to true if the semester offerings say they are offered.
                            if sem_offer_args[-1] == "Semester" or \
                                    ("S1" in sem_offer_value and "S2" in sem_offer_value):
                                s1_offer = True
                                s2_offer = True
                            elif sem_offer_args[-1] == "S1":
                                s1_offer = True
                            elif sem_offer_args[-1] == "S2":
                                s2_offer = True

                            # Assume all courses are worth 6 units,
                            # since unit value is not included in the CASS course list excel sheet.
                            row_data = [row[col_index["Course Code"]].value, year,
                                        row[col_index["Course Title"]].value, 6, s1_offer, s2_offer]
                            uploaded_file.append(row_data)

            # If the uploaded excel sheet is in custom format specified in line 23,
            # then simply convert from the excel format to a list of lists.
            else:
                for row in sheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    uploaded_file.append(row_data)

        else:
            # Reading the '%' using the csv import module came from:
            # https://stackoverflow.com/questions/13992971/reading-and-parsing-a-tsv-file-then-manipulating-it-for-saving-as-csv-efficie

            # % is used instead of comma since the course name may include commas (which would break this function)
            uploaded_file = csv.reader(uploaded_file, delimiter='%')

        # Stores the index of the column containing the data type of each row,
        # so that the right data is stored in the right column
        # This would also allow columns to be in any order, and courses/subplans would still be added.
        map = {}
        for row in uploaded_file:
            if first_row_checked:
                if content_type == 'Courses':
                    # If number of columns from file doesn't match the model, return error to user.
                    if len(row) != 6:
                        any_error = True
                        break

                    course_instance = CourseModel()
                    course_instance.code = row[map['code']]
                    course_instance.year = int(row[map['year']])
                    course_instance.name = row[map['name']]
                    course_instance.units = int(row[map['units']])
                    course_instance.offeredSem1 = bool(row[map['offeredSem1']])
                    course_instance.offeredSem2 = bool(row[map['offeredSem2']])
                    course_str = course_instance.code + " - " + course_instance.name

                    # Save the course instance
                    try:
                        course_instance.save()
                        any_success = True
                        correctly_uploaded.append(course_str)
                    except:
                        error_message = course_str + " Couldn't add: Check for duplicate course"
                        failed_to_upload.append(error_message)
                        any_error = True

                elif content_type == 'Subplans':
                    if len(row) != 5:
                        any_error = True
                        break

                    subplan_instance = SubplanModel()
                    subplan_instance.code = row[map['code']]
                    subplan_instance.year = int(row[map['year']])
                    subplan_instance.name = row[map['name']]
                    subplan_instance.units = int(row[map['units']])
                    subplan_instance.planType = str(row[map['planType']])
                    subplan_str = str(subplan_instance.year) + " - " + subplan_instance.code + " - " + \
                        subplan_instance.name

                    # Save the subplan instance
                    try:
                        subplan_instance.save()
                        any_success = True
                        correctly_uploaded.append(subplan_str)
                    except:
                        any_error = True
                        failed_to_upload.append(subplan_str)

            else:
                i = 0
                for col in row:
                    map[col] = i
                    i += 1
                first_row_checked = True

        # Display error messages depending on the level of success of bulk upload.
        # There are 3 categories: All successful, some successful or none successful.
        if any_success and not any_error:
            context['user_msg'] = "All items has been added successfully!"
            context['err_type'] = "success"

        elif any_success and any_error:
            # for course in failed_to_upload:
            failed_str = ""
            correct_str = ""

            for record in failed_to_upload:
                failed_str = failed_str + "<br> - " + record

            for record in correctly_uploaded:
                correct_str = correct_str + "<br> - " + record

            context['user_msg'] = "Some items could not be added. They may already be present in the database" \
                                  " or the data may not be in the correct format. Please check the " \
                                  + content_type + \
                                  " list and try manually adding ones that failed through the dedicated " \
                                  "forms. <br><br>The following " + content_type + " could not be added: " + \
                                  failed_str + "<br><br> " \
                                  "The following " + content_type + " uploaded successfully: " + correct_str
            context['err_type'] = "warn"

        elif not any_success and any_error:
            context['user_msg'] = "All items failed to be added. " \
                                  "Either you have already uploaded the same contents, " \
                                  "or the format of the file is incorrect. Please try again."
            context['err_type'] = "error"

    return render(request, 'bulkupload.html', context=context)
